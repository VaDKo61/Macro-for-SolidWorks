import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Side, Border, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from functions.archive.create_drill_sheet import get_ready


def marge_format_excel(names_files: list) -> Workbook:
    """Marge and format excel for Energocomfort"""

    title_name = names_files[0].split(".")[0].replace('_', '/')
    title_sheet = {
        'sheet1': f'{title_name} (Основное оборудование)',
        'sheet2': f'{title_name} (Механическая обработка)'
    }
    if len(names_files) == 2:
        title_sheet['sheet3'] = f'{title_name} (Рама)'
    else:
        for number in range(1, len(names_files) + 1):
            title_sheet[f'sheet{number + 2}'] = f'{title_name} (Рама часть № {number})'

    # Load basic Excel
    wb_basic = load_workbook(filename=r'\\192.168.1.14\SolidWorks\Excel SW\ТЗ' + f'\{names_files[0]}')
    ws = wb_basic.active

    # Edit basic data Excel in list
    ws.insert_rows(0)
    sort_data = sort_by_name(ws)
    sort_data = merge_identical_positions(sort_data)
    title = first_second_row(ws)
    if get_ready():
        return
    sort_data = sort_by_fourth(sort_data, title)
    sort_data = number_column(sort_data)

    # Add frame in data
    if len(names_files) == 2:
        sort_data = edit_add_frame(names_files[1], sort_data, title)
    else:
        for name_frame in names_files[1:]:
            sort_data = edit_add_frame(name_frame, sort_data, title)

    # Delete and add data in Excel
    ws.delete_rows(3, ws.max_row)
    for row in sort_data:
        ws.append(row)

    # Past title in table
    count = 1
    for column in ws.iter_rows(min_col=1, max_col=1):
        if column[0].value is None:
            column[0].value = title_sheet[f'sheet{count}']
            count += 1

    # Summ mass
    for column in ws.iter_rows(min_col=9, max_col=10, max_row=1):
        column[0].value = 'Общий вес:'
        column[1].value = summ_mass(sort_data)

    # Set width column
    name_size_column = {
        'A': 5,
        'B': 16.5,
        'C': 80,
        'D': 13.5,
        'E': 10.4,
        'F': 11.8,
        'G': 25,
        'H': 22.5,
        'I': 21.3,
        'J': 11,
        'K': 16.5
    }
    for name_column, size in name_size_column.items():
        ws.column_dimensions[f'{name_column}'].width = size

    # Set height column
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20

    # Create Style
    normal_text = NamedStyle(name='normal_text')
    normal_text.font = Font(name='ISOCPEUR', italic=True, size=14)
    bd = Side(style='thin', color='000000')
    normal_text.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal_text.alignment = Alignment(horizontal='center', vertical='center')
    wb_basic.add_named_style(normal_text)

    bold_normal_text = NamedStyle(name='bold_normal_text')
    bold_normal_text.font = Font(name='ISOCPEUR', bold=True, italic=True, size=15)
    bold_normal_text.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    bold_normal_text.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb_basic.add_named_style(bold_normal_text)

    normal_float = NamedStyle(name='normal_float')
    normal_float.font = Font(name='ISOCPEUR', italic=True, size=14)
    bd = Side(style='thin', color='000000')
    normal_float.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal_float.alignment = Alignment(horizontal='center', vertical='center')
    normal_float.number_format = '0.0'
    wb_basic.add_named_style(normal_float)

    bold_title_text = NamedStyle(name='bold_title_text')
    bold_title_text.font = Font(name='ISOCPEUR', bold=True, italic=True, size=18)
    bold_title_text.alignment = Alignment(horizontal='left', vertical='center')
    wb_basic.add_named_style(bold_title_text)

    # Set Style
    for row in ws.iter_rows():
        if row[0].value == '№':
            for cell in row:
                cell.style = bold_normal_text
            ws.row_dimensions[row[0].row].height = 38
        elif isinstance(row[0].value, str):
            for cell in row:
                cell.style = bold_title_text
            ws.row_dimensions[row[0].row].height = 25
        else:
            for cell in row:
                cell.style = normal_text
    for cells in ws.iter_cols(min_col=10, max_col=10):
        for cell in cells:
            if not isinstance(cell.value, str) and cell.value is not None:
                cell.style = normal_float

    return wb_basic


def sort_by_name(work_sheet: Worksheet) -> list:
    """Return sorted list by column 'C' (Name)"""

    sort_list = []
    for row_sort in work_sheet.iter_rows(min_row=3, values_only=True):
        if row_sort[2] not in [None, '-']:
            sort_list.append(row_sort)
    sort_list.sort(key=lambda x: x[2])
    return sort_list


def merge_identical_positions(sort_data: list) -> list:
    """Return merge list by identical position"""
    new_data = []
    delete_list = []
    for i in range(len(sort_data)):
        current_line = sort_data[i]
        if current_line in delete_list:
            continue
        for check_line in sort_data[i + 1:]:
            if current_line[2] == check_line[2]:
                if current_line[1] == check_line[1] and current_line[3] == check_line[3]:
                    current_line = list(current_line)
                    current_line[4] = current_line[4] + check_line[4]
                    delete_list.append(tuple(check_line))
                else:
                    continue
        new_data.append(current_line)
    return new_data


def first_second_row(work_sheet: Worksheet) -> list:
    """Return firs row"""
    title = []
    for two_row in work_sheet.iter_rows(min_row=1, max_row=2, values_only=True):
        title.append(two_row)
    return title


def sort_by_fourth(sorted_data: list, title: list) -> list:
    """Row, with int in fourth column, go end"""
    first_list = []
    second_list = []
    for row in sorted_data:
        if row[3] is None:
            first_list.append(row)
        else:
            second_list.append(row)
    return first_list + title + second_list


def number_column(sorted_data: list) -> list:
    """Number column A (№) and past '-' in clear cell"""
    count = 1
    count_row = 0
    for row in sorted_data:
        if isinstance(row[0], int):
            row = list(row)
            row[0] = count
            if row[3] is None:
                row[3], row[10] = '-', '-'
            if row[10] is None:
                row[10] = '-'
            sorted_data[count_row] = tuple(row)
            count += 1
            count_row += 1
        else:
            count = 1
            count_row += 1
    return sorted_data


def edit_add_frame(names: list, sort_data: list, title: list) -> list:
    """Edit and add in data frame material"""
    # Load frame Excel
    wb_frame = load_workbook(filename=r'\\192.168.1.14\SolidWorks\Excel SW\ТЗ' + f'\{names}')
    ws_frame = wb_frame.active
    ws_frame.insert_rows(0)
    for row in ws_frame.iter_rows(min_col=11, max_col=11):  # Add 11 column
        for cell in row:
            cell.value = '-'

    # Edit frame data Excel in list
    sort_data_frame = sort_by_name(ws_frame)
    sort_data_frame = merge_identical_positions(sort_data_frame)
    sort_data_frame = number_column(sort_data_frame)

    # Unite sort_data and sort_data_frame
    sort_data.append(title[0])
    sort_data.append(title[1])
    sort_data += sort_data_frame
    return sort_data


def summ_mass(sort_data):
    result = []
    for i in sort_data:
        if isinstance(i[9], float) or isinstance(i[9], int):
            result.append(i[4] * i[9])
        elif isinstance(i[9], str):
            if i[9].replace('.', '').replace(',', '').isdigit():
                result.append(i[4] * float(i[9].replace(',', '.')))
    return str(round(sum(result) * 1.15, 1))


def conversion_excel():
    # Search file in derictory
    name_files = ''
    for root, dirs, files in os.walk(r"\\192.168.1.14\SolidWorks\Excel SW\ТЗ"):
        name_files = files

    # Validate file in derictory
    for name in name_files:
        if 'xlsx' not in name:
            raise FileExistsError('Файлы с другим расширением')

    # Group name by: what need unite
    group_names = []
    for name_file in name_files:
        if 'рам' not in name_file.lower():
            group_names.append([name_file])
    for name_file in name_files:
        if 'рам' in name_file.lower():
            name = name_file.split(' ')[0]
            for group in range(len(group_names)):
                if name in group_names[group][0]:
                    group_names[group].append(name_file)
    for name_files in group_names:
        wb = marge_format_excel(name_files)
        wb.save(r'\\192.168.1.14\SolidWorks\Excel SW' + '\\' + f'{name_files[0].split(".")[0]} Итог.xls')
        print('Excel успешно преобразован.')
